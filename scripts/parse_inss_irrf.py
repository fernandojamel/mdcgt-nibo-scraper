#!/usr/bin/env python3
"""
Parser do "RATEIO DE IMPOSTOS SOBRE FOLHA PAGTO." (Excel) — aba
"COMPOSIÇÃO DA DARF". DARF de folha paga pela Matriz, rateada por loja.

Separa:
  - IRRF  = linha "0561-07 - IRRF ..." (por loja)
  - INSS  = "TOTAL DARF 1410" − IRRF (líquido, já com as deduções)
  (INSS + IRRF == TOTAL DARF por loja)

Colunas por CNPJ no cabeçalho: /0001-22 = Tijuca, /0002-03 = Metropolitano,
/0003-94 = Taquara (resíduo somado no Tijuca/Matriz).

Competência: do nome do arquivo (RATEIO - ABRIL - 2026 - ...) ou passada
explicitamente (na recorrência o arquivo vem com nome temporário).

Uso:
    python parse_inss_irrf.py <xlsx_path> [--pretty]
"""

import sys
import os
import json
import re
import openpyxl

MESES = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARCO': 3, 'MARÇO': 3, 'ABRIL': 4,
    'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9,
    'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
}


def num(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def loja_do_cnpj(suf):
    return 'Metropolitano' if suf == '0002-03' else 'Tijuca'  # 0001-22 e 0003-94 -> Tijuca


def comp_do_nome(nome):
    up = nome.upper()
    ano = re.search(r'(20\d{2})', up)
    for mnome, mnum in MESES.items():
        if mnome in up and ano:
            return f'{ano.group(1)}-{mnum:02d}-01'
    return None


def parse_inss_irrf(path, competencia=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    nome_aba = next((s for s in wb.sheetnames if 'DARF' in s.upper()), wb.sheetnames[0])
    ws = wb[nome_aba]
    rows = list(ws.iter_rows(values_only=True))

    # Mapa coluna -> loja, pelo CNPJ no cabeçalho.
    col_loja = {}
    for r in rows[:6]:
        for j, c in enumerate(r):
            if c and 'CNPJ' in str(c):
                m = re.search(r'/(\d{4}-\d{2})', str(c))
                if m:
                    col_loja[j] = loja_do_cnpj(m.group(1))

    def acha_linha(pred):
        for r in rows:
            if r and r[0] and pred(str(r[0])):
                return r
        return None

    linha_irrf = acha_linha(lambda s: '0561' in s)
    linha_darf = acha_linha(lambda s: 'TOTAL DARF' in s.upper())

    por_loja = {'Tijuca': {'INSS': 0.0, 'IRRF': 0.0},
                'Metropolitano': {'INSS': 0.0, 'IRRF': 0.0}}
    for j, loja in col_loja.items():
        irrf_v = num(linha_irrf[j]) if linha_irrf and j < len(linha_irrf) else 0.0
        darf_v = num(linha_darf[j]) if linha_darf and j < len(linha_darf) else 0.0
        por_loja[loja]['IRRF'] += irrf_v
        por_loja[loja]['INSS'] += darf_v - irrf_v

    total_darf = sum(num(linha_darf[j]) for j in col_loja) if linha_darf else 0.0
    soma = sum(v['INSS'] + v['IRRF'] for v in por_loja.values())
    # Só é válido se achou a linha TOTAL DARF e as colunas por loja (senão é
    # outro arquivo que o filtro amplo pescou) — evita ingerir zeros.
    reconcilia = (linha_darf is not None and bool(col_loja)
                  and total_darf > 0 and abs(soma - total_darf) < 0.02)

    comp = competencia
    if comp and len(comp) == 7:  # 'AAAA-MM'
        comp = comp + '-01'
    if not comp:
        comp = comp_do_nome(os.path.basename(path))

    return {
        'competencia': comp,
        'imposto': 'INSS_IRRF',
        'total_darf': round(total_darf, 2),
        'lojas': {k: {'INSS': round(v['INSS'], 2), 'IRRF': round(v['IRRF'], 2)}
                  for k, v in por_loja.items()},
        'reconcilia': reconcilia,
    }


def main():
    if len(sys.argv) < 2:
        print('Uso: parse_inss_irrf.py <xlsx_path> [--pretty]', file=sys.stderr)
        sys.exit(1)
    pretty = '--pretty' in sys.argv
    r = parse_inss_irrf(sys.argv[1])
    print(json.dumps(r, indent=2 if pretty else None, ensure_ascii=False))


if __name__ == '__main__':
    main()
